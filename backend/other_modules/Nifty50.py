
import os
import sys
import time
import pandas as pd
import mysql.connector
import yfinance as yf
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Path setup (mirrors backend_alert.py) ────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)

try:
    from auth_manager import get_client
except ImportError:
    print(f"❌ auth_manager.py not found in: {PROJECT_ROOT}")
    sys.exit(1)

# ── DB config ─────────────────────────────────────────────────────────────────
DB_CONFIG = dict(
    host="localhost", user="root", password="Root",
    database="stocks", port=3306, connect_timeout=10
)

# ── Nifty 50 — (display name, NSE symbol for 5paisa, yfinance .NS, sector) ───
NIFTY50 = [
    ("Adani Enterprises",    "ADANIENT",    "ADANIENT.NS",    "Diversified"),
    ("Adani Ports",          "ADANIPORTS",  "ADANIPORTS.NS",  "Infrastructure"),
    ("Apollo Hospitals",     "APOLLOHOSP",  "APOLLOHOSP.NS",  "Healthcare"),
    ("Asian Paints",         "ASIANPAINT",  "ASIANPAINT.NS",  "Paints / Consumer"),
    ("Axis Bank",            "AXISBANK",    "AXISBANK.NS",    "Banking"),
    ("Bajaj Auto",           "BAJAJ-AUTO",  "BAJAJ-AUTO.NS",  "Automobiles"),
    ("Bajaj Finance",        "BAJFINANCE",  "BAJFINANCE.NS",  "NBFC / Finance"),
    ("Bajaj Finserv",        "BAJAJFINSV",  "BAJAJFINSV.NS",  "NBFC / Finance"),
    ("BPCL",                 "BPCL",        "BPCL.NS",        "Oil & Gas"),
    ("Bharti Airtel",        "BHARTIARTL",  "BHARTIARTL.NS",  "Telecom"),
    ("Britannia",            "BRITANNIA",   "BRITANNIA.NS",   "FMCG / Food"),
    ("Cipla",                "CIPLA",       "CIPLA.NS",       "Pharmaceuticals"),
    ("Coal India",           "COALINDIA",   "COALINDIA.NS",   "Mining / Coal"),
    ("Divi's Labs",          "DIVISLAB",    "DIVISLAB.NS",    "Pharmaceuticals"),
    ("Dr Reddy's",           "DRREDDY",     "DRREDDY.NS",     "Pharmaceuticals"),
    ("Eicher Motors",        "EICHERMOT",   "EICHERMOT.NS",   "Automobiles"),
    ("Grasim Industries",    "GRASIM",      "GRASIM.NS",      "Diversified"),
    ("HCL Technologies",     "HCLTECH",     "HCLTECH.NS",     "IT / Technology"),
    ("HDFC Bank",            "HDFCBANK",    "HDFCBANK.NS",    "Banking"),
    ("HDFC Life",            "HDFCLIFE",    "HDFCLIFE.NS",    "Insurance"),
    ("Hero MotoCorp",        "HEROMOTOCO",  "HEROMOTOCO.NS",  "Automobiles"),
    ("Hindalco Industries",  "HINDALCO",    "HINDALCO.NS",    "Metals"),
    ("Hindustan Unilever",   "HINDUNILVR",  "HINDUNILVR.NS",  "FMCG / Consumer"),
    ("ICICI Bank",           "ICICIBANK",   "ICICIBANK.NS",   "Banking"),
    ("IndusInd Bank",        "INDUSINDBK",  "INDUSINDBK.NS",  "Banking"),
    ("Infosys",              "INFY",        "INFY.NS",        "IT / Technology"),
    ("ITC",                  "ITC",         "ITC.NS",         "FMCG / Tobacco"),
    ("JSW Steel",            "JSWSTEEL",    "JSWSTEEL.NS",    "Metals"),
    ("Kotak Mahindra Bank",  "KOTAKBANK",   "KOTAKBANK.NS",   "Banking"),
    ("L&T",                  "LT",          "LT.NS",          "Infrastructure"),
    ("M&M",                  "M&M",         "M&M.NS",         "Automobiles"),
    ("Maruti Suzuki",        "MARUTI",      "MARUTI.NS",      "Automobiles"),
    ("NTPC",                 "NTPC",        "NTPC.NS",        "Power Generation"),
    ("Nestle India",         "NESTLEIND",   "NESTLEIND.NS",   "FMCG / Food"),
    ("ONGC",                 "ONGC",        "ONGC.NS",        "Oil & Gas"),
    ("Power Grid Corp",      "POWERGRID",   "POWERGRID.NS",   "Power Transmission"),
    ("Reliance Industries",  "RELIANCE",    "RELIANCE.NS",    "Oil & Gas / Conglomerate"),
    ("SBI Life Insurance",   "SBILIFE",     "SBILIFE.NS",     "Insurance"),
    ("Shriram Finance",      "SHRIRAMFIN",  "SHRIRAMFIN.NS",  "NBFC / Finance"),
    ("State Bank of India",  "SBIN",        "SBIN.NS",        "Banking"),
    ("Sun Pharma",           "SUNPHARMA",   "SUNPHARMA.NS",   "Pharmaceuticals"),
    ("TCS",                  "TCS",         "TCS.NS",         "IT / Technology"),
    ("Tata Consumer",        "TATACONSUM",  "TATACONSUM.NS",  "FMCG / Food"),
    ("Tata Motors",          "TATAMOTORS",  "TATAMOTORS.NS",  "Automobiles"),
    ("Tata Steel",           "TATASTEEL",   "TATASTEEL.NS",   "Metals"),
    ("Tech Mahindra",        "TECHM",       "TECHM.NS",       "IT / Technology"),
    ("Titan Company",        "TITAN",       "TITAN.NS",       "Jewellery / Consumer"),
    ("UltraTech Cement",     "ULTRACEMCO",  "ULTRACEMCO.NS",  "Cement"),
    ("Wipro",                "WIPRO",       "WIPRO.NS",       "IT / Technology"),
    ("Eternal Ltd (Zomato)", "ETERNAL",     "ETERNAL.NS",     "Food Tech / Consumer"),
]

# ── Industry Avg PE benchmarks ────────────────────────────────────────────────
SECTOR_AVG_PE = {
    "Banking":                     "15–18",
    "NBFC / Finance":              "25–30",
    "IT / Technology":             "22–26",
    "Information Technology":      "22–26",
    "Pharmaceuticals":             "28–35",
    "Healthcare":                  "60–80",
    "FMCG / Consumer":             "55–65",
    "FMCG / Food":                 "55–65",
    "FMCG / Tobacco":              "15–18",
    "FMCG":                        "45–65",
    "Automobiles":                 "20–30",
    "Automobile":                  "20–30",
    "Auto Components":             "18–25",
    "Metals":                      "16–18",
    "Metals & Mining":             "16–18",
    "Mining / Coal":               "12–14",
    "Oil & Gas":                   "10–14",
    "Oil, Gas & Consumable Fuels": "10–14",
    "Oil & Gas / Conglomerate":    "20–25",
    "Power Generation":            "18–22",
    "Power Transmission":          "22–26",
    "Power":                       "18–22",
    "Infrastructure":              "25–35",
    "Construction":                "25–35",
    "Cement":                      "35–45",
    "Insurance":                   "55–70",
    "Financial Services":          "20–30",
    "Telecom":                     "35–50",
    "Telecommunication":           "35–50",
    "Paints / Consumer":           "50–65",
    "Jewellery / Consumer":        "80–100",
    "Consumer Durables":           "40–60",
    "Consumer Services":           "40–60",
    "Realty":                      "30–50",
    "Chemicals":                   "25–35",
    "Diversified":                 "20–30",
    "Services":                    "25–40",
    "Capital Goods":               "30–45",
    "Food Tech / Consumer":        "200–300",
    "Media & Entertainment":       "30–50",
    "Forest Materials":            "15–25",
}

# ═══════════════════════════════════════════════════════════════
#  SCRIP MASTER
# ═══════════════════════════════════════════════════════════════

def load_scrip_master():
    csv_path = os.path.join(PROJECT_ROOT, 'ScripMaster_all.csv')
    try:
        df    = pd.read_csv(csv_path)
        df_eq = df[df['Series'] == 'EQ'].copy() if 'Series' in df.columns else df.copy()
        by_symbol, by_name = {}, {}
        for _, r in df_eq.iterrows():
            code = int(r['ScripCode'])
            if 'ScripName' in df_eq.columns:
                by_symbol[str(r['ScripName']).strip().upper()] = code
            by_name[str(r['Name']).strip().upper()] = code
        print(f"✅ ScripMaster: {len(df_eq)} EQ records")
        return by_name, by_symbol, df_eq
    except FileNotFoundError:
        print(f"❌ ScripMaster_all.csv not found at: {csv_path}")
        sys.exit(1)


def get_scrip_code(symbol, by_name, by_symbol, df):
    key = symbol.strip().upper()

    def _meta(code):
        row = df[df['ScripCode'] == code]
        if row.empty:
            return code, 'N', 'C'
        r = row.iloc[0]
        return code, str(r.get('Exch','N')).strip().upper(), str(r.get('ExchType','C')).strip().upper()

    if key in by_symbol: return _meta(by_symbol[key])
    if key in by_name:   return _meta(by_name[key])

    hit = df[df['Name'].str.upper().str.contains(key, na=False)]
    if hit.empty and 'ScripName' in df.columns:
        hit = df[df['ScripName'].str.upper().str.contains(key, na=False)]
    if not hit.empty:
        r = hit.iloc[0]
        return int(r['ScripCode']), str(r.get('Exch','N')).strip().upper(), str(r.get('ExchType','C')).strip().upper()

    return None, None, None

# ═══════════════════════════════════════════════════════════════
#  5PAISA — Price / 52W / 3W change
# ═══════════════════════════════════════════════════════════════

def fetch_5paisa_prices(client, scrip_code, symbol, exch, exch_type):
    result = dict(current_price=None, high_52w=None, low_52w=None, price_change_3w=None)
    try:
        end     = datetime.now().strftime('%Y-%m-%d')
        s_52w   = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
        s_3w    = (datetime.now() - timedelta(days=28)).strftime('%Y-%m-%d')

        df = client.historical_data(exch, exch_type, int(scrip_code), '1d', s_52w, end)
        if df is None or df.empty:
            return result

        df['Datetime'] = pd.to_datetime(df['Datetime'])
        df = df.sort_values('Datetime')

        result['current_price'] = round(float(df.iloc[-1]['Close']), 2)
        result['high_52w']      = round(float(df['High'].max()), 2)
        result['low_52w']       = round(float(df['Low'].min()), 2)

        df3w = df[df['Datetime'] >= pd.Timestamp(s_3w)]
        if not df3w.empty:
            p3 = float(df3w.iloc[0]['Close'])
            if p3 > 0:
                pct = ((result['current_price'] - p3) / p3) * 100
                result['price_change_3w'] = f"{pct:+.1f}%"
    except Exception as e:
        print(f"    ⚠ 5paisa error [{symbol}]: {e}")
    return result

# ═══════════════════════════════════════════════════════════════
#  YFINANCE — PE + sector + price fallback
# ═══════════════════════════════════════════════════════════════

def fetch_yf_data(yf_symbol):
    """
    Returns dict: pe, sector, current_price, high_52w, low_52w, price_change_3w
    Same trailingPE logic as your comparison.py fetch_yfinance_data().
    """
    out = dict(pe=None, sector=None, current_price=None,
               high_52w=None, low_52w=None, price_change_3w=None)
    try:
        ticker = yf.Ticker(yf_symbol)
        info   = ticker.info

        # PE — same as comparison.py
        pe = info.get('trailingPE') or info.get('forwardPE')
        out['pe'] = round(float(pe), 1) if pe and float(pe) > 0 else None

        # Sector
        out['sector'] = info.get('sector') or info.get('industry')

        # Price data (used as fallback if 5paisa fails)
        cp = info.get('currentPrice') or info.get('regularMarketPrice')
        if cp:
            out['current_price'] = round(float(cp), 2)
            out['high_52w']      = round(float(info.get('fiftyTwoWeekHigh', 0)), 2) or None
            out['low_52w']       = round(float(info.get('fiftyTwoWeekLow',  0)), 2) or None

            hist = ticker.history(period="1mo")
            if not hist.empty:
                p3 = float(hist['Close'].iloc[0])
                if p3 > 0:
                    pct = ((float(cp) - p3) / p3) * 100
                    out['price_change_3w'] = f"{pct:+.1f}%"
    except Exception as e:
        print(f"    ⚠ yfinance error [{yf_symbol}]: {e}")
    return out

# ═══════════════════════════════════════════════════════════════
#  BUILD ONE ROW
# ═══════════════════════════════════════════════════════════════

def build_row(name, sym5p, yf_sym, sector, client, by_name, by_symbol, csv_df):
    # Step 1 — yfinance (PE always from here, prices as fallback)
    yf = fetch_yf_data(yf_sym)
    time.sleep(0.2)

    # Step 2 — 5paisa prices (overrides yfinance prices if available)
    prices = dict(current_price=None, high_52w=None, low_52w=None, price_change_3w=None)
    scrip_code, exch, exch_type = get_scrip_code(sym5p, by_name, by_symbol, csv_df)
    if scrip_code and client:
        prices = fetch_5paisa_prices(client, scrip_code, sym5p, exch, exch_type)
    else:
        print(f"    ⚠ {sym5p} not in ScripMaster — using yfinance prices")

    # Step 3 — merge: prefer 5paisa for prices, yfinance for PE/sector
    cur   = prices['current_price']   or yf['current_price']
    h52   = prices['high_52w']        or yf['high_52w']
    l52   = prices['low_52w']         or yf['low_52w']
    ch3w  = prices['price_change_3w'] or yf['price_change_3w']
    pe    = yf['pe']
    res_sector = (sector if sector and sector.strip() not in ("", "Unknown")
                  else (yf['sector'] or "Unknown"))

    return {
        "Stock":              name,
        "Sector":             res_sector,
        "Current Price (₹)":  cur   if cur   is not None else "N/A",
        "52-Week High (₹)":   h52   if h52   is not None else "N/A",
        "52-Week Low (₹)":    l52   if l52   is not None else "N/A",
        "PE Ratio":           pe    if pe    is not None else "N/A",
        "Industry Avg PE":    SECTOR_AVG_PE.get(res_sector, "N/A"),
        "Price Change (~3W)": ch3w  if ch3w  is not None else "N/A",
    }

# ═══════════════════════════════════════════════════════════════
#  DATABASE
# ═══════════════════════════════════════════════════════════════

def get_db_stocks():
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cur  = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT stock_id, symbol, company_name, sector, industry
            FROM stocks_master WHERE is_active = 1 ORDER BY symbol
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        print(f"✅ DB: {len(rows)} active stocks")
        return rows
    except Exception as e:
        print(f"❌ DB error: {e}")
        return []

# ═══════════════════════════════════════════════════════════════
#  EXCEL
# ═══════════════════════════════════════════════════════════════

COLS       = ["Stock", "Sector", "Current Price (₹)", "52-Week High (₹)",
              "52-Week Low (₹)", "PE Ratio", "Industry Avg PE", "Price Change (~3W)"]
COL_WIDTHS = [24, 26, 18, 18, 16, 12, 16, 20]
HDR_FILL   = PatternFill("solid", fgColor="1F3864")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ALT_FILL   = PatternFill("solid", fgColor="EEF2FF")
WHT_FILL   = PatternFill("solid", fgColor="FFFFFF")
_T         = Side(style="thin", color="CCCCCC")
BDR        = Border(left=_T, right=_T, top=_T, bottom=_T)


def write_sheet(ws, title, rows):
    ws.title = title
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for ri, row in enumerate(rows, 2):
        fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
        for ci, col in enumerate(COLS, 1):
            val = row.get(col, "")
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = fill; c.border = BDR
            c.alignment = Alignment(horizontal="left" if ci <= 2 else "center", vertical="center")
            if col == "Price Change (~3W)" and isinstance(val, str) and val not in ("N/A",""):
                try:
                    pct = float(val.replace("%","").replace("+",""))
                    c.font = Font(name="Arial", size=10, bold=True,
                                  color="006100" if pct >= 0 else "9C0006")
                except: pass
        ws.row_dimensions[ri].height = 20

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    note = ws.cell(row=len(rows)+3, column=1,
                   value=f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}  |  "
                         f"Price/52W/3W: 5paisa API (yfinance fallback)  |  PE: Yahoo Finance")
    note.font = Font(name="Arial", size=9, italic=True, color="888888")

# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("   Nifty50 + DB Screener  |  5paisa + yfinance")
    print("=" * 60)

    client = get_client()
    if not client:
        print("⚠ 5paisa client unavailable — yfinance will be used for all prices")

    by_name, by_symbol, csv_df = load_scrip_master()

    # Tab 1
    print(f"\n[Tab 1] Nifty 50 ({len(NIFTY50)} stocks)")
    tab1 = []
    for i, (name, sym5p, yf_sym, sector) in enumerate(NIFTY50, 1):
        print(f"  [{i:02d}/{len(NIFTY50)}] {name}")
        tab1.append(build_row(name, sym5p, yf_sym, sector, client, by_name, by_symbol, csv_df))
        time.sleep(0.3)

    # Tab 2
    db_stocks = get_db_stocks()
    print(f"\n[Tab 2] DB Stocks ({len(db_stocks)} stocks)")
    tab2 = []
    for i, s in enumerate(db_stocks, 1):
        sym    = s['symbol']
        name   = s.get('company_name') or sym
        sector = s.get('sector') or s.get('industry') or ""
        yf_sym = f"{sym}.NS"
        print(f"  [{i:03d}/{len(db_stocks)}] {name} ({sym})")
        tab2.append(build_row(name, sym, yf_sym, sector, client, by_name, by_symbol, csv_df))
        time.sleep(0.3)

    # Write Excel
    print("\nWriting Excel...")
    wb = Workbook()
    write_sheet(wb.active, "Nifty 50", tab1)
    write_sheet(wb.create_sheet("My DB Stocks"), "My DB Stocks", tab2)

    out = os.path.join(SCRIPT_DIR, "Nifty50_5paisa_Screener.xlsx")
    wb.save(out)
    print(f"\n✅ Saved: {out}")
    print(f"   Tab 1 (Nifty 50):  {len(tab1)} rows")
    print(f"   Tab 2 (DB Stocks): {len(tab2)} rows")
    print("=" * 60)


if __name__ == "__main__":
    main()